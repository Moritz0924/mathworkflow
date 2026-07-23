from __future__ import annotations

import math
import shutil
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .entities import BombEvaluation, BombPlan


WORKSPACE_ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_BY_QUESTION = {
    "Q3": WORKSPACE_ROOT / "03_data" / "raw" / "result1.xlsx",
    "Q4": WORKSPACE_ROOT / "03_data" / "raw" / "result2.xlsx",
    "Q5": WORKSPACE_ROOT / "03_data" / "raw" / "result3.xlsx",
}


def _write_used_fields(
    worksheet,
    row: int,
    plan: BombPlan,
    evaluation: BombEvaluation,
    *,
    heading_column: int,
    speed_column: int,
    release_column: int,
    detonation_column: int,
    duration_column: int,
) -> None:
    release = evaluation.release_point_xyz
    detonation = evaluation.detonation_point_xyz
    if release is None or detonation is None:
        raise ValueError("used plan must have a physical evaluation")
    worksheet.cell(row, heading_column).value = math.degrees(plan.heading_rad) % 360.0
    worksheet.cell(row, speed_column).value = plan.speed_mps
    for offset, value in enumerate(release):
        worksheet.cell(row, release_column + offset).value = value
    for offset, value in enumerate(detonation):
        worksheet.cell(row, detonation_column + offset).value = value
    worksheet.cell(row, duration_column).value = evaluation.effective_duration_s


def _clear_calculated_fields(worksheet, row: int, columns: Iterable[int]) -> None:
    for column in columns:
        worksheet.cell(row, column).value = None


def export_candidate_template(
    question_id: str,
    plans: Iterable[BombPlan],
    evaluations: Iterable[BombEvaluation],
    destination: Path,
) -> Path:
    if question_id not in TEMPLATE_BY_QUESTION:
        raise ValueError("only Q3, Q4, and Q5 have official Excel templates")
    template = TEMPLATE_BY_QUESTION[question_id]
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, destination)
    worksheet = load_workbook(destination).active
    plan_list = list(plans)
    evaluation_by_key = {(item.plan.uav_id, item.plan.bomb_id): item for item in evaluations}

    if question_id == "Q3":
        for plan in plan_list:
            row = plan.bomb_id + 1
            if not plan.is_used:
                _clear_calculated_fields(worksheet, row, range(1, 11))
                worksheet.cell(row, 3).value = plan.bomb_id
                continue
            _write_used_fields(
                worksheet,
                row,
                plan,
                evaluation_by_key[(plan.uav_id, plan.bomb_id)],
                heading_column=1,
                speed_column=2,
                release_column=4,
                detonation_column=7,
                duration_column=10,
            )
            worksheet.cell(row, 3).value = plan.bomb_id
    elif question_id == "Q4":
        rows = {str(worksheet.cell(row, 1).value): row for row in range(2, 5)}
        for plan in plan_list:
            row = rows[plan.uav_id]
            if not plan.is_used:
                _clear_calculated_fields(worksheet, row, range(2, 11))
                continue
            _write_used_fields(
                worksheet,
                row,
                plan,
                evaluation_by_key[(plan.uav_id, plan.bomb_id)],
                heading_column=2,
                speed_column=3,
                release_column=4,
                detonation_column=7,
                duration_column=10,
            )
    else:
        rows = {
            (str(worksheet.cell(row, 1).value), int(worksheet.cell(row, 4).value)): row
            for row in range(2, 17)
        }
        for plan in plan_list:
            row = rows[(plan.uav_id, plan.bomb_id)]
            if not plan.is_used:
                _clear_calculated_fields(worksheet, row, (2, 3, 5, 6, 7, 8, 9, 10, 11, 12))
                continue
            _write_used_fields(
                worksheet,
                row,
                plan,
                evaluation_by_key[(plan.uav_id, plan.bomb_id)],
                heading_column=2,
                speed_column=3,
                release_column=5,
                detonation_column=8,
                duration_column=11,
            )
            worksheet.cell(row, 12).value = plan.missile_id

    worksheet.parent.save(destination)
    return destination
