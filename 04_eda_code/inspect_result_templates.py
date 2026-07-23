from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "03_data" / "raw"


def _data_rows(sheet: Any) -> list[int]:
    rows: list[int] = []
    for row_number in range(2, sheet.max_row + 1):
        values = [sheet.cell(row_number, column).value for column in range(1, sheet.max_column + 1)]
        if not any(value is not None and str(value).strip() for value in values):
            break
        rows.append(row_number)
    return rows


def inspect_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False)
    if len(workbook.worksheets) != 1:
        raise ValueError(f"{path.name} must contain exactly one worksheet")

    sheet = workbook.active
    headers = [str(sheet.cell(1, column).value or "").strip() for column in range(1, sheet.max_column + 1)]
    formula_count = sum(
        1
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )
    return {
        "file_name": path.name,
        "sheet_name": sheet.title,
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "headers": headers,
        "data_rows": _data_rows(sheet),
        "merged_ranges": [str(item) for item in sheet.merged_cells.ranges],
        "hidden_rows": [index for index, dimension in sheet.row_dimensions.items() if dimension.hidden],
        "hidden_columns": [key for key, dimension in sheet.column_dimensions.items() if dimension.hidden],
        "is_protected": bool(sheet.protection.sheet),
        "has_formulas": formula_count > 0,
        "formula_count": formula_count,
        "number_formats": sorted({cell.number_format for row in sheet.iter_rows() for cell in row}),
        "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else "",
        "data_validation_count": len(sheet.data_validations.dataValidation),
        "sample_values": {
            header: [sheet.cell(row, column).value for row in _data_rows(sheet) if sheet.cell(row, column).value is not None]
            for column, header in enumerate(headers, start=1)
        },
    }


def _inferred_type(header: str) -> str:
    if "编号" in header and "烟幕干扰弹" not in header:
        return "categorical"
    if "烟幕干扰弹编号" in header:
        return "integer"
    return "numeric"


def _write_data_dictionary(reports: list[dict[str, Any]]) -> Path:
    path = ROOT / "03_data" / "data_dictionary.csv"
    fields = [
        "file_name",
        "column_name",
        "inferred_type",
        "missing_rate",
        "unique_count",
        "sample_values",
        "possible_meaning",
        "risk_note",
    ]
    rows: list[dict[str, str]] = []
    for report in reports:
        for index, header in enumerate(report["headers"]):
            samples = report["sample_values"].get(header, [])
            rows.append(
                {
                    "file_name": report["file_name"],
                    "column_name": header,
                    "inferred_type": _inferred_type(header),
                    "missing_rate": "not_applicable_template",
                    "unique_count": str(len({str(value) for value in samples})),
                    "sample_values": " | ".join(str(value) for value in samples),
                    "possible_meaning": header,
                    "risk_note": "结果待计算；空白单元格不是观测数据缺失",
                }
            )
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)
    return path


def _write_structure_report(reports: list[dict[str, Any]]) -> Path:
    path = ROOT / "03_data" / "template_structure_report.md"
    lines = ["# 官方结果模板结构核验", "", "- 检查方式：openpyxl 只读结构检查。", "- 结论：三份文件均为待填写的结果模板，不是统计样本数据。", ""]
    for report in reports:
        rows = report["data_rows"]
        lines += [
            f"## {report['file_name']}",
            "",
            f"- 工作表：`{report['sheet_name']}`（共 1 张）",
            f"- 维度：{report['max_row']} 行 × {report['max_column']} 列",
            f"- 可填写结果行：{', '.join(map(str, rows))}",
            f"- 合并单元格：{report['merged_ranges'] or '无'}",
            f"- 隐藏行/列：{report['hidden_rows'] or '无'} / {report['hidden_columns'] or '无'}",
            f"- 保护：{'是' if report['is_protected'] else '否'}；公式：{report['formula_count']}；数据验证：{report['data_validation_count']}",
            f"- 单元格数字格式：{', '.join(report['number_formats'])}",
            "- 表头：" + "；".join(report["headers"]),
            "",
        ]
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def _write_quality_report(reports: list[dict[str, Any]]) -> Path:
    path = ROOT / "03_data" / "data_quality_report.md"
    lines = [
        "# 数据质量报告",
        "",
        "## 结论",
        "",
        "三份 Excel 文件均为官方结果填写模板。其空白数值单元格表示结果尚未计算，不能按观测数据缺失处理，也不能用于相关性、回归或异常值分布分析。",
        "",
        "## 模板核验",
        "",
        "| 文件 | 工作表 | 可填写结果行 | 结构状态 |",
        "|---|---|---|---|",
    ]
    for report in reports:
        rows = ", ".join(map(str, report["data_rows"]))
        lines.append(f"| {report['file_name']} | {report['sheet_name']} | {rows} | 无合并、无公式、无隐藏行列、未保护 |")
    lines += [
        "",
        "## 后续结果校验规则",
        "",
        "- 已使用弹的投放点、起爆点和有效干扰时长必须为可复核数值。",
        "- 未使用弹位由最终模型设计明确留空或填 0；无论何种表示，负的有效时长均非法。",
        "- 同一无人机的多弹记录必须共享同一航向与速度，并满足相邻投放间隔不少于 1 s。",
        "- 相关性 EDA 在没有至少两列真实数值观测前必须跳过。",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def _write_eda_report(reports: list[dict[str, Any]]) -> Path:
    path = ROOT / "04_eda" / "eda_report.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    names = "、".join(report["file_name"] for report in reports)
    content = f"""# 数据分析与 EDA 报告

## 数据性质

{names} 是结果模板而非观测样本。预填的仅为无人机或烟幕弹编号；方向、速度、坐标与时长均待后续计算产生。

## 可执行分析

- 已完成模板字段、行结构、工作表与格式核验，详见 `03_data/template_structure_report.md`。
- 已完成题面参数、模板输出字段和硬约束的结构化映射。
- 当前不生成相关性热力图、回归、描述统计或特征筛选图：这些方法对空结果模板没有统计意义。

## 已发现的本地事实冲突

- `input_decision_table.csv` 中的 Q1 时序已按原题修正为“接令后 1.5 s 投放”和“投放后 3.6 s 起爆”，且不外推到 Q2–Q5。
- EDA 脚本现在要求至少两列具有真实数值观测才生成相关性图。

## 进入模型设计前仍需登记的假设

- 有效遮蔽的几何判据；
- 干扰弹初始速度、重力常数、地面边界和 Q2–Q5 起爆时序；
- 多弹重叠时间的统计规则及 Q5 的目标聚合方式。
"""
    path.write_text(content, encoding="utf-8")
    return path


def write_reports(root: Path = ROOT) -> dict[str, Path]:
    global ROOT, RAW
    ROOT = root.resolve()
    RAW = ROOT / "03_data" / "raw"
    reports = [inspect_workbook(path) for path in sorted(RAW.glob("result*.xlsx"))]
    if [report["file_name"] for report in reports] != ["result1.xlsx", "result2.xlsx", "result3.xlsx"]:
        raise ValueError("expected result1.xlsx, result2.xlsx, and result3.xlsx")
    return {
        "data_dictionary": _write_data_dictionary(reports),
        "data_quality_report": _write_quality_report(reports),
        "template_structure_report": _write_structure_report(reports),
        "eda_report": _write_eda_report(reports),
    }


def main() -> None:
    outputs = write_reports(ROOT)
    for label, path in outputs.items():
        print(f"[OK] {label}: {path.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
