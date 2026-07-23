from __future__ import annotations

import math
import json
import sys
import tempfile
import unittest
from pathlib import Path


CODE_ROOT = Path(__file__).resolve().parents[1] / "06_code"
if str(CODE_ROOT) not in sys.path:
    sys.path.insert(0, str(CODE_ROOT))


class DomainObjectTests(unittest.TestCase):
    def test_q1_plan_uses_the_official_fixed_timing(self) -> None:
        from src.constants import MISSILES, UAVS
        from src.entities import BombPlan

        plan = BombPlan(
            uav_id="FY1",
            bomb_id=1,
            missile_id="M1",
            is_used=True,
            heading_rad=math.pi,
            speed_mps=120.0,
            release_time_s=1.5,
            delay_s=3.6,
        )

        self.assertEqual((20000.0, 0.0, 2000.0), MISSILES["M1"].initial_position)
        self.assertEqual((17800.0, 0.0, 1800.0), UAVS["FY1"].initial_position)
        self.assertEqual(5.1, plan.detonation_time_s)


class DynamicsTests(unittest.TestCase):
    def test_q1_ballistic_detonation_point_is_derived_from_motion(self) -> None:
        from src.dynamics import detonation_point, missile_hit_time, missile_position, smoke_valid_end_time, uav_position
        from src.entities import BombPlan

        plan = BombPlan("FY1", 1, "M1", True, math.pi, 120.0, 1.5, 3.6)

        self.assertEqual((17620.0, 0.0, 1800.0), tuple(uav_position("FY1", math.pi, 120.0, 1.5)))
        self.assertEqual((17188.0, 0.0, 1736.496), tuple(detonation_point(plan)))
        self.assertAlmostEqual(66.999170807, missile_hit_time("M1"), places=6)
        self.assertAlmostEqual(0.0, float(missile_position("M1", missile_hit_time("M1"))[0]), places=6)
        self.assertEqual(25.1, smoke_valid_end_time(plan, "M1"))

    def test_zero_horizontal_velocity_sensitivity_keeps_detonation_above_release_point(self) -> None:
        from src.dynamics import detonation_point, release_point
        from src.entities import BombPlan

        plan = BombPlan("FY1", 1, "M1", True, math.pi, 120.0, 1.5, 3.6)
        release = release_point(plan)
        inherited = detonation_point(plan)
        no_horizontal_velocity = detonation_point(plan, horizontal_velocity_factor=0.0)

        self.assertAlmostEqual(release[0], no_horizontal_velocity[0], places=9)
        self.assertAlmostEqual(release[1], no_horizontal_velocity[1], places=9)
        self.assertNotAlmostEqual(inherited[0], no_horizontal_velocity[0], places=6)

    def test_ground_hold_sensitivity_does_not_end_smoke_at_ground_contact(self) -> None:
        from src.dynamics import smoke_valid_end_time
        from src.entities import BombPlan

        plan = BombPlan("FY3", 1, "M1", True, math.pi, 120.0, 0.0, 11.9)
        terminated = smoke_valid_end_time(plan, "M1")
        held = smoke_valid_end_time(plan, "M1", smoke_ground_mode="hold")

        self.assertGreater(held, terminated)


class GeometryTests(unittest.TestCase):
    def test_line_of_sight_uses_the_segment_and_closed_smoke_boundary(self) -> None:
        from src.occlusion import fraction_points_occluded, is_all_points_occluded, line_of_sight_clearance
        from src.target_geometry import target_sample_points

        points = target_sample_points(angular_samples=4, height_samples=2, radial_samples=1)
        self.assertIn((7.0, 200.0, 0.0), {tuple(point) for point in points})
        self.assertIn((0.0, 200.0, 10.0), {tuple(point) for point in points})

        line_lambda, clearance = line_of_sight_clearance((0, 0, 0), (0, 10, 0), (10, 5, 0))
        self.assertEqual(0.5, line_lambda)
        self.assertEqual(10.0, clearance)
        self.assertTrue(is_all_points_occluded((0, 0, 0), [(0, 10, 0)], (10, 5, 0), 10.0))
        self.assertFalse(is_all_points_occluded((0, 0, 0), [(0, 10, 0)], (0, -1, 0), 10.0))
        self.assertEqual(
            0.5,
            fraction_points_occluded((0, 0, 0), [(0, 10, 0), (20, 10, 0)], (0, 5, 0), 1.0),
        )


class IntervalTests(unittest.TestCase):
    def test_continuous_interval_scan_refines_boundaries_and_merges_overlap(self) -> None:
        from src.intervals import find_true_intervals, find_true_intervals_fixed_grid, interval_duration, merge_intervals

        intervals = find_true_intervals(lambda time_s: 1.0 <= time_s <= 2.0, 0.0, 3.0, scan_step_s=0.4, tolerance_s=1e-7)

        self.assertEqual(1, len(intervals))
        self.assertAlmostEqual(1.0, intervals[0][0], places=5)
        self.assertAlmostEqual(2.0, intervals[0][1], places=5)
        self.assertAlmostEqual(1.0, interval_duration(intervals), places=5)
        self.assertEqual([(0.0, 3.0)], merge_intervals([(0.0, 1.0), (0.9, 2.0), (2.00001, 3.0)], tolerance_s=1e-4))

        fixed_grid = find_true_intervals_fixed_grid(
            lambda time_s: 1.0 <= time_s <= 2.0,
            0.0,
            3.0,
            scan_step_s=0.5,
        )
        self.assertEqual([(1.0, 2.0)], fixed_grid)


class ConstraintTests(unittest.TestCase):
    def test_q3_rejects_release_gap_smaller_than_one_second(self) -> None:
        from src.constraints import validate_plans
        from src.entities import BombPlan

        plans = [
            BombPlan("FY1", 1, "M1", True, math.pi, 120.0, 1.0, 1.0),
            BombPlan("FY1", 2, "M1", True, math.pi, 120.0, 1.5, 1.0),
            BombPlan("FY1", 3, "M1", True, math.pi, 120.0, 3.0, 1.0),
        ]

        violations = validate_plans("Q3", plans)

        self.assertIn("release_gap", {violation.code for violation in violations})


class EvaluatorTests(unittest.TestCase):
    def test_q1_evaluator_keeps_fixed_decision_and_returns_physical_time_window(self) -> None:
        from src.evaluator import EvaluatorConfig, evaluate_q1

        evaluation = evaluate_q1(EvaluatorConfig(angular_samples=6, height_samples=2, radial_samples=1, scan_step_s=0.25))

        self.assertEqual("M1", evaluation.plan.missile_id)
        self.assertEqual(1.5, evaluation.plan.release_time_s)
        self.assertEqual(3.6, evaluation.plan.delay_s)
        self.assertGreaterEqual(evaluation.effective_duration_s, 0.0)
        self.assertLessEqual(evaluation.smoke_valid_end_s, 25.1)
        self.assertTrue(all(5.1 <= start <= end <= evaluation.smoke_valid_end_s for start, end in evaluation.occlusion_intervals))

    def test_fixed_grid_baseline_uses_unrefined_scan_boundaries(self) -> None:
        from src.evaluator import EvaluatorConfig, evaluate_q1

        evaluation = evaluate_q1(
            EvaluatorConfig(
                angular_samples=6,
                height_samples=2,
                radial_samples=1,
                scan_step_s=0.25,
                interval_method="fixed_grid",
            )
        )

        self.assertAlmostEqual(evaluation.effective_duration_s / 0.25, round(evaluation.effective_duration_s / 0.25), places=8)

    def test_coverage_ratio_sensitivity_can_relax_all_point_requirement(self) -> None:
        from src.evaluator import EvaluatorConfig, evaluate_q1

        evaluation = evaluate_q1(
            EvaluatorConfig(
                angular_samples=6,
                height_samples=2,
                radial_samples=1,
                scan_step_s=0.25,
                min_coverage_fraction=0.0,
            )
        )

        self.assertAlmostEqual(20.0, evaluation.effective_duration_s, places=8)


class SolverTests(unittest.TestCase):
    def test_q2_solver_returns_a_kinematically_valid_single_bomb_candidate(self) -> None:
        from src.constraints import validate_plans
        from src.evaluator import EvaluatorConfig, evaluate_q1
        from src.solvers import SolverConfig, solve_q2

        solution = solve_q2(
            SolverConfig(
                evaluator=EvaluatorConfig(angular_samples=6, height_samples=2, radial_samples=1, scan_step_s=0.25),
                differential_evolution_maxiter=1,
                differential_evolution_popsize=3,
                local_refinement=False,
                random_seed=20250701,
            )
        )

        self.assertEqual("Q2", solution.question_id)
        self.assertEqual(1, len(solution.plans))
        self.assertEqual([], validate_plans("Q2", solution.plans))
        self.assertEqual("M1", solution.plans[0].missile_id)
        self.assertGreaterEqual(solution.primary_durations_s["M1"], evaluate_q1(EvaluatorConfig(angular_samples=6, height_samples=2, radial_samples=1, scan_step_s=0.25)).effective_duration_s)

    def test_q3_solver_keeps_shared_flight_variables_and_release_order(self) -> None:
        from src.constraints import validate_plans
        from src.evaluator import EvaluatorConfig, evaluate_q1
        from src.solvers import SolverConfig, solve_q2, solve_q3

        config = SolverConfig(
            evaluator=EvaluatorConfig(angular_samples=6, height_samples=2, radial_samples=1, scan_step_s=0.25),
            differential_evolution_maxiter=1,
            differential_evolution_popsize=3,
            local_refinement=False,
            random_seed=20250702,
        )
        solution = solve_q3(config)

        self.assertEqual("Q3", solution.question_id)
        self.assertEqual(3, len(solution.plans))
        self.assertEqual([], validate_plans("Q3", solution.plans))
        self.assertEqual(1, len({plan.heading_rad for plan in solution.plans}))
        self.assertEqual(1, len({plan.speed_mps for plan in solution.plans}))
        self.assertTrue(all(later.release_time_s - earlier.release_time_s >= 1.0 for earlier, later in zip(solution.plans, solution.plans[1:])))
        self.assertGreaterEqual(solution.primary_durations_s["M1"], evaluate_q1(EvaluatorConfig(angular_samples=6, height_samples=2, radial_samples=1, scan_step_s=0.25)).effective_duration_s)
        self.assertGreaterEqual(solution.primary_durations_s["M1"], solve_q2(config).primary_durations_s["M1"])

    def test_q4_solver_uses_each_required_uav_once(self) -> None:
        from src.constraints import validate_plans
        from src.evaluator import EvaluatorConfig, evaluate_q1
        from src.solvers import SolverConfig, solve_q2, solve_q4

        config = SolverConfig(
            evaluator=EvaluatorConfig(angular_samples=6, height_samples=2, radial_samples=1, scan_step_s=0.25),
            differential_evolution_maxiter=1,
            differential_evolution_popsize=3,
            local_refinement=False,
            random_seed=20250703,
        )
        solution = solve_q4(config)

        self.assertEqual("Q4", solution.question_id)
        self.assertEqual({"FY1", "FY2", "FY3"}, {plan.uav_id for plan in solution.plans})
        self.assertEqual([], validate_plans("Q4", solution.plans))
        self.assertGreaterEqual(solution.primary_durations_s["M1"], evaluate_q1(EvaluatorConfig(angular_samples=6, height_samples=2, radial_samples=1, scan_step_s=0.25)).effective_duration_s)
        self.assertGreaterEqual(solution.primary_durations_s["M1"], solve_q2(config).primary_durations_s["M1"])

    def test_q5_solver_returns_all_fifteen_slots_and_reports_three_target_feasibility(self) -> None:
        from src.constraints import validate_plans
        from src.evaluator import EvaluatorConfig
        from src.solvers import SolverConfig, solve_q5

        solution = solve_q5(
            SolverConfig(
                evaluator=EvaluatorConfig(angular_samples=6, height_samples=2, radial_samples=1, scan_step_s=0.25),
                differential_evolution_maxiter=0,
                differential_evolution_popsize=2,
                local_refinement=False,
                random_seed=20250704,
            )
        )

        self.assertEqual("Q5", solution.question_id)
        self.assertEqual(15, len(solution.plans))
        self.assertEqual([], validate_plans("Q5", solution.plans))
        self.assertEqual({"M1", "M2", "M3"}, set(solution.primary_durations_s))
        self.assertTrue(solution.diagnostics["feasible"])
        self.assertTrue(all(duration > 0.0 for duration in solution.primary_durations_s.values()))
        self.assertEqual(15, len(solution.diagnostics["score_matrix"]))
        self.assertEqual(
            {"uav_id", "missile_id", "effective_duration_s"},
            set(solution.diagnostics["score_matrix"][0]),
        )


class ExporterTests(unittest.TestCase):
    def test_q5_export_copies_template_and_leaves_unused_calculated_fields_blank(self) -> None:
        from openpyxl import load_workbook

        from src.evaluator import EvaluatorConfig
        from src.exporters import export_candidate_template
        from src.solvers import SolverConfig, solve_q5

        source = Path(__file__).resolve().parents[1] / "03_data" / "raw" / "result3.xlsx"
        source_before = source.read_bytes()
        solution = solve_q5(
            SolverConfig(
                evaluator=EvaluatorConfig(angular_samples=6, height_samples=2, radial_samples=1, scan_step_s=0.25),
                differential_evolution_maxiter=0,
                differential_evolution_popsize=2,
                local_refinement=False,
                random_seed=20250705,
            )
        )

        with tempfile.TemporaryDirectory() as raw:
            destination = Path(raw) / "result3_candidate.xlsx"
            export_candidate_template("Q5", solution.plans, solution.evaluations, destination)
            exported = load_workbook(destination).active

            self.assertTrue(destination.exists())
            unused_plan = next(plan for plan in solution.plans if not plan.is_used)
            row = next(index for index in range(2, 17) if exported.cell(index, 1).value == unused_plan.uav_id and exported.cell(index, 4).value == unused_plan.bomb_id)
            self.assertIsNone(exported.cell(row, 2).value)
            self.assertIsNone(exported.cell(row, 11).value)
            self.assertIsNone(exported.cell(row, 12).value)
        self.assertEqual(source_before, source.read_bytes())

    def test_q3_export_preserves_bomb_identifier_column(self) -> None:
        from openpyxl import load_workbook

        from src.evaluator import EvaluatorConfig
        from src.exporters import export_candidate_template
        from src.solvers import SolverConfig, solve_q3

        solution = solve_q3(
            SolverConfig(
                evaluator=EvaluatorConfig(angular_samples=6, height_samples=2, radial_samples=1, scan_step_s=0.25),
                differential_evolution_maxiter=0,
                differential_evolution_popsize=2,
                local_refinement=False,
                random_seed=20250706,
            )
        )
        with tempfile.TemporaryDirectory() as raw:
            destination = Path(raw) / "result1_candidate.xlsx"
            export_candidate_template("Q3", solution.plans, solution.evaluations, destination)
            exported = load_workbook(destination).active

            self.assertEqual([1, 2, 3], [exported.cell(row, 3).value for row in range(2, 5)])
            self.assertAlmostEqual(solution.evaluations[0].release_point_xyz[0], exported.cell(2, 4).value)


class EntrypointTests(unittest.TestCase):
    def test_pipeline_can_skip_candidate_template_exports_for_numeric_validation(self) -> None:
        from src.evaluator import EvaluatorConfig
        from src.solvers import SolverConfig
        from run_all import run_pipeline

        with tempfile.TemporaryDirectory() as raw:
            output_dir = Path(raw)
            run_pipeline(
                output_dir=output_dir,
                solver_config=SolverConfig(
                    evaluator=EvaluatorConfig(angular_samples=6, height_samples=2, radial_samples=1, scan_step_s=0.25),
                    differential_evolution_maxiter=0,
                    differential_evolution_popsize=2,
                    local_refinement=False,
                    random_seed=20250800,
                ),
                export_templates=False,
            )

            self.assertTrue((output_dir / "metrics_summary.csv").exists())
            self.assertFalse((output_dir / "result1_candidate.xlsx").exists())

    def test_quick_pipeline_writes_candidate_artifacts_for_all_five_questions(self) -> None:
        from src.evaluator import EvaluatorConfig
        from src.solvers import SolverConfig
        from run_all import run_pipeline

        with tempfile.TemporaryDirectory() as raw:
            output_dir = Path(raw)
            validation_config = SolverConfig(
                evaluator=EvaluatorConfig(angular_samples=6, height_samples=2, radial_samples=1, scan_step_s=0.25),
                differential_evolution_maxiter=0,
                differential_evolution_popsize=2,
                local_refinement=False,
                random_seed=20250799,
            )
            manifest = run_pipeline(quick=True, output_dir=output_dir, solver_config=validation_config)

            self.assertEqual("candidate", manifest["result_status"])
            self.assertEqual({"Q1", "Q2", "Q3", "Q4", "Q5"}, set(manifest["questions"]))
            self.assertTrue((output_dir / "candidate_run_manifest.json").exists())
            self.assertTrue((output_dir / "q5_results.csv").exists())
            self.assertTrue((output_dir / "result3_candidate.xlsx").exists())
            self.assertEqual("candidate", json.loads((output_dir / "candidate_run_manifest.json").read_text(encoding="utf-8"))["result_status"])
            self.assertEqual(20250799, manifest["solver_config"]["random_seed"])


if __name__ == "__main__":
    unittest.main()
