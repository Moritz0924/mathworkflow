from __future__ import annotations

import csv
import json
import re
import subprocess
import sys
import zipfile
from collections import Counter
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

ROOT = Path(__file__).resolve().parents[1]
PROBLEM_INBOX = ROOT / "00_problem" / "inbox"
DATA_RAW = ROOT / "03_data" / "raw"
OUT_PROBLEM = ROOT / "00_problem" / "problem_statement.md"
OUT_ATTACH = ROOT / "00_problem" / "attachments_overview.md"
OUT_PROFILE = ROOT / "01_task_analysis" / "problem_model_profile.csv"
OUT_DECOMP = ROOT / "01_task_analysis" / "task_decomposition.md"
OUT_MISSING = ROOT / "01_task_analysis" / "missing_information.md"
OUT_DATA_DICT = ROOT / "03_data" / "data_dictionary.csv"
OUT_DATA_REPORT = ROOT / "03_data" / "data_quality_report.md"
OUT_MODEL_ROUTE = ROOT / "05_model" / "model_route.md"
OUT_SUMMARY = ROOT / "01_task_analysis" / "auto_start_summary.md"

SUPPORTED_TEXT = {".txt", ".md"}
SUPPORTED_DOCX = {".docx"}
SUPPORTED_PDF = {".pdf"}
SUPPORTED_DATA = {".csv", ".tsv", ".xlsx", ".xls", ".json", ".geojson"}
IMAGE_EXT = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tif", ".tiff"}

PROBLEM_TYPES = ["数据评价型", "预测回归型", "优化决策型", "机理仿真型", "机器学习型", "综合开放型"]
MODEL_FAMILIES = ["统计评价", "预测回归", "优化决策", "机理仿真", "机器学习", "空间网络", "综合集成"]

KEYWORD_RULES = [
    {
        "problem_type": "优化决策型",
        "model_family": "优化决策",
        "keywords": ["最优", "优化", "规划", "约束", "分配", "调度", "路径", "成本最小", "收益最大", "选址", "排班", "运输", "容量", "资源配置"],
        "baseline": "贪心策略/线性规划松弛/枚举搜索",
        "risk": "目标函数、约束条件和可行性解释不足",
    },
    {
        "problem_type": "数据评价型",
        "model_family": "统计评价",
        "keywords": ["评价", "评估", "指标", "权重", "排序", "排名", "综合得分", "优劣", "等级", "打分", "TOPSIS", "层次分析", "熵权"],
        "baseline": "等权评价/TOPSIS/熵权法",
        "risk": "指标含义、权重来源和排序稳定性不足",
    },
    {
        "problem_type": "预测回归型",
        "model_family": "预测回归",
        "keywords": ["预测", "趋势", "未来", "回归", "拟合", "时间序列", "销量", "增长", "衰减", "需求", "误差", "残差", "外推"],
        "baseline": "线性回归/ARIMA/指数平滑",
        "risk": "误差诊断、外推边界和变量解释不足",
    },
    {
        "problem_type": "机理仿真型",
        "model_family": "机理仿真",
        "keywords": ["机理", "仿真", "模拟", "微分方程", "动力学", "传播", "演化", "状态转移", "守恒", "参数", "情景", "ODE", "PDE"],
        "baseline": "差分仿真/简化微分方程模型",
        "risk": "假设、参数来源和敏感性验证不足",
    },
    {
        "problem_type": "机器学习型",
        "model_family": "机器学习",
        "keywords": ["分类", "识别", "聚类", "训练", "测试集", "机器学习", "神经网络", "随机森林", "SVM", "特征", "准确率", "召回率", "泛化"],
        "baseline": "逻辑回归/随机森林/LightGBM",
        "risk": "训练验证划分、过拟合控制和可解释性不足",
    },
    {
        "problem_type": "综合开放型",
        "model_family": "空间网络",
        "keywords": ["空间", "地图", "经纬度", "地理", "网络", "节点", "边", "流量", "连通", "社区", "路径", "拓扑", "GIS"],
        "baseline": "网络指标/空间统计/最短路径",
        "risk": "空间尺度、节点含义和流向解释不足",
    },
]


def safe_read_text(path: Path) -> str:
    for enc in ("utf-8-sig", "utf-8", "gb18030", "gbk", "latin1"):
        try:
            return path.read_text(encoding=enc)
        except UnicodeDecodeError:
            continue
        except Exception:
            break
    return ""


def read_docx_text(path: Path) -> str:
    try:
        with zipfile.ZipFile(path) as z:
            xml = z.read("word/document.xml").decode("utf-8", errors="ignore")
        xml = re.sub(r"</w:p>", "\n", xml)
        xml = re.sub(r"<[^>]+>", "", xml)
        return xml.replace("&lt;", "<").replace("&gt;", ">").replace("&amp;", "&")
    except Exception as exc:
        return f"[DOCX 抽取失败: {exc}]"


def read_pdf_text(path: Path) -> str:
    try:
        try:
            from pypdf import PdfReader  # type: ignore
        except Exception:
            from PyPDF2 import PdfReader  # type: ignore
        reader = PdfReader(str(path))
        pages = []
        for page in reader.pages[:12]:
            pages.append(page.extract_text() or "")
        return "\n".join(pages)
    except Exception as exc:
        return f"[PDF 未自动抽取正文：{exc}。建议安装 pypdf，或同时放入 txt/md 题目文本。]"


def collect_problem_files() -> Tuple[str, List[Dict[str, str]]]:
    PROBLEM_INBOX.mkdir(parents=True, exist_ok=True)
    items = []
    parts = []
    for path in sorted(PROBLEM_INBOX.iterdir()):
        if path.name.startswith(".") or path.name.lower() == "readme.md" or not path.is_file():
            continue
        ext = path.suffix.lower()
        text = ""
        status = "已登记"
        if ext in SUPPORTED_TEXT:
            text = safe_read_text(path)
            status = "已抽取文本"
        elif ext in SUPPORTED_DOCX:
            text = read_docx_text(path)
            status = "已抽取文本"
        elif ext in SUPPORTED_PDF:
            text = read_pdf_text(path)
            status = "尝试抽取 PDF 文本"
        elif ext in IMAGE_EXT:
            status = "图片已登记；未 OCR，建议补充文字版题目"
        else:
            status = "未知题目格式，仅登记"
        items.append({"file": path.name, "ext": ext, "status": status, "chars": str(len(text))})
        if text.strip():
            parts.append(f"\n\n## 来源：{path.name}\n\n{text.strip()}\n")
    return "\n".join(parts).strip(), items


def split_questions(problem_text: str) -> List[Tuple[str, str]]:
    text = problem_text.strip()
    if not text:
        return [("Q1", "未能自动读取题目文本，请补充 txt/md 题目文本。"), ("Q2", "待识别"), ("Q3", "待识别")]
    patterns = [
        r"(?=(?:问题\s*[一二三四五六七八九十0-9]+|第\s*[一二三四五六七八九十0-9]+\s*问|任务\s*[一二三四五六七八九十0-9]+)[:：、\s])",
        r"(?=\n\s*[（(]?[一二三四五六七八九十0-9]+[）).、]\s*)",
    ]
    chunks = []
    for pat in patterns:
        raw = [c.strip() for c in re.split(pat, "\n" + text) if c.strip()]
        candidates = [c for c in raw if re.search(r"问题|第\s*[一二三四五六七八九十0-9]+\s*问|任务", c[:80])]
        if len(candidates) >= 2:
            chunks = candidates
            break
    if not chunks:
        # Fallback: detect paragraphs containing question-style verbs.
        paras = [p.strip() for p in re.split(r"\n\s*\n+", text) if p.strip()]
        target = [p for p in paras if re.search(r"建立|求解|评价|预测|优化|分析|给出|设计|判断|计算", p)]
        chunks = target[:4] if len(target) >= 2 else [text[:1500]]
    out = []
    for i, c in enumerate(chunks[:6], start=1):
        c = re.sub(r"\s+", " ", c).strip()
        out.append((f"Q{i}", c[:1200]))
    return out


def score_rule(text: str, data_hints: str = "") -> Dict[str, str]:
    full = f"{text}\n{data_hints}".lower()
    scores = []
    for rule in KEYWORD_RULES:
        score = 0
        hit_words = []
        for kw in rule["keywords"]:
            if kw.lower() in full:
                score += 1
                hit_words.append(kw)
        scores.append((score, len(hit_words), rule, hit_words))
    scores.sort(key=lambda x: (x[0], x[1]), reverse=True)
    best_score, _, best, hit_words = scores[0]
    if best_score == 0:
        best = {
            "problem_type": "综合开放型",
            "model_family": "综合集成",
            "baseline": "描述统计/基础模型对比",
            "risk": "题型不清，需人工复核模型路线",
        }
        hit_words = []
    main_model = infer_main_model(best["model_family"], full)
    return {
        "problem_type": best["problem_type"],
        "model_family": best["model_family"],
        "main_model": main_model,
        "baseline_model": best["baseline"],
        "risk_focus": best["risk"],
        "confidence_0_1": confidence(best_score),
        "notes": "命中关键词：" + ("、".join(hit_words[:8]) if hit_words else "不足，按综合题处理"),
    }


def confidence(score: int) -> str:
    if score >= 5:
        return "0.86"
    if score >= 3:
        return "0.72"
    if score >= 1:
        return "0.58"
    return "0.42"


def infer_main_model(family: str, text: str) -> str:
    table = {
        "优化决策": [
            ("遗传算法/模拟退火", ["遗传", "模拟退火", "全局搜索"]),
            ("整数规划/混合整数规划", ["整数", "0-1", "选址", "排班"]),
            ("多目标规划 + Pareto 分析", ["多目标", "pareto", "权衡"]),
            ("线性/非线性规划", ["线性规划", "非线性", "约束"]),
        ],
        "统计评价": [
            ("熵权-TOPSIS/CRITIC-TOPSIS", ["客观权重", "熵权", "topsis"]),
            ("层次分析法 AHP", ["层次", "专家", "判断矩阵"]),
            ("模糊综合评价", ["模糊", "等级", "隶属"]),
        ],
        "预测回归": [
            ("时间序列预测", ["时间序列", "arima", "周期", "季节"]),
            ("多元回归/正则化回归", ["回归", "变量", "影响因素"]),
            ("灰色预测/指数平滑", ["小样本", "灰色", "指数平滑"]),
        ],
        "机理仿真": [
            ("微分方程动力学模型", ["微分", "动力学", "传播"]),
            ("蒙特卡洛/离散事件仿真", ["随机", "蒙特卡洛", "排队", "事件"]),
        ],
        "机器学习": [
            ("随机森林/梯度提升树", ["特征", "重要性", "分类", "回归"]),
            ("聚类模型", ["聚类", "分群", "类别"]),
            ("神经网络", ["神经网络", "深度", "lstm", "cnn"]),
        ],
        "空间网络": [
            ("空间统计 + 网络分析", ["网络", "节点", "边", "经纬度", "空间"]),
        ],
    }
    for name, kws in table.get(family, []):
        if any(k.lower() in text for k in kws):
            return name
    defaults = {
        "优化决策": "约束优化模型",
        "统计评价": "指标评价模型",
        "预测回归": "预测回归模型",
        "机理仿真": "机理仿真模型",
        "机器学习": "机器学习对比模型",
        "空间网络": "空间网络模型",
        "综合集成": "多模型综合集成",
    }
    return defaults.get(family, "综合建模方案")


def analyze_csv(path: Path, delimiter: str = ",") -> Dict[str, object]:
    rows = 0
    headers: List[str] = []
    sample_values: Dict[str, List[str]] = {}
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            text = f.read(4096)
            if path.suffix.lower() == ".tsv":
                delimiter = "\t"
            else:
                try:
                    delimiter = csv.Sniffer().sniff(text).delimiter
                except Exception:
                    delimiter = ","
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f, delimiter=delimiter)
            for i, row in enumerate(reader):
                if i == 0:
                    headers = [str(x).strip() or f"col_{j+1}" for j, x in enumerate(row)]
                    sample_values = {h: [] for h in headers}
                    continue
                rows += 1
                if i <= 6:
                    for h, v in zip(headers, row):
                        if len(sample_values[h]) < 5 and str(v).strip():
                            sample_values[h].append(str(v).strip()[:40])
    except UnicodeDecodeError:
        with path.open("r", encoding="gb18030", newline="") as f:
            reader = csv.reader(f, delimiter=delimiter)
            for i, row in enumerate(reader):
                if i == 0:
                    headers = [str(x).strip() or f"col_{j+1}" for j, x in enumerate(row)]
                    sample_values = {h: [] for h in headers}
                    continue
                rows += 1
    except Exception as exc:
        return {"file": path.name, "type": path.suffix.lower(), "status": f"读取失败: {exc}", "rows": "", "columns": "", "headers": []}
    return {"file": path.name, "type": path.suffix.lower(), "status": "已读取", "rows": rows, "columns": len(headers), "headers": headers, "samples": sample_values}


def analyze_xlsx(path: Path) -> Dict[str, object]:
    try:
        from openpyxl import load_workbook  # type: ignore
        wb = load_workbook(path, read_only=True, data_only=True)
        sheets = []
        headers_all = []
        for ws in wb.worksheets[:8]:
            rows = ws.max_row or 0
            cols = ws.max_column or 0
            header = []
            if rows >= 1:
                for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
                    if cell is not None:
                        header.append(str(cell).strip())
            sheets.append({"sheet": ws.title, "rows": rows, "columns": cols, "headers": header[:30]})
            headers_all.extend(header)
        return {"file": path.name, "type": path.suffix.lower(), "status": "已读取", "sheets": sheets, "headers": headers_all[:80]}
    except Exception as exc:
        return {"file": path.name, "type": path.suffix.lower(), "status": f"读取失败: {exc}", "headers": []}


def analyze_json(path: Path) -> Dict[str, object]:
    try:
        data = json.loads(safe_read_text(path))
        keys = []
        if isinstance(data, dict):
            keys = list(data.keys())[:50]
        elif isinstance(data, list) and data and isinstance(data[0], dict):
            keys = list(data[0].keys())[:50]
        return {"file": path.name, "type": path.suffix.lower(), "status": "已读取", "keys": keys, "headers": keys}
    except Exception as exc:
        return {"file": path.name, "type": path.suffix.lower(), "status": f"读取失败: {exc}", "headers": []}


def collect_data_files() -> List[Dict[str, object]]:
    DATA_RAW.mkdir(parents=True, exist_ok=True)
    rows = []
    for path in sorted(DATA_RAW.iterdir()):
        if path.name.startswith(".") or not path.is_file():
            continue
        ext = path.suffix.lower()
        if ext in {".csv", ".tsv"}:
            rows.append(analyze_csv(path))
        elif ext in {".xlsx", ".xls"}:
            rows.append(analyze_xlsx(path))
        elif ext in {".json", ".geojson"}:
            rows.append(analyze_json(path))
        else:
            rows.append({"file": path.name, "type": ext, "status": "已登记；未自动解析", "headers": []})
    return rows


def data_hints(data_rows: List[Dict[str, object]]) -> str:
    parts = []
    for d in data_rows:
        parts.append(str(d.get("file", "")))
        for key in ("headers", "keys"):
            val = d.get(key)
            if isinstance(val, list):
                parts.extend(map(str, val[:40]))
        for sheet in d.get("sheets", []) if isinstance(d.get("sheets"), list) else []:
            parts.extend(map(str, sheet.get("headers", [])[:30]))
    return " ".join(parts)


def infer_data_shape(data_rows: List[Dict[str, object]], qtext: str) -> str:
    hint = data_hints(data_rows).lower() + " " + qtext.lower()
    if any(k in hint for k in ["经度", "纬度", "lon", "lat", "lng", "坐标", "location"]):
        return "空间/地理数据"
    if any(k in hint for k in ["日期", "时间", "year", "month", "day", "time", "date"]):
        return "时间序列或面板数据"
    if any(k in hint for k in ["类别", "标签", "class", "label", "类型"]):
        return "分类标签数据"
    if any(k in hint for k in ["指标", "得分", "评分", "权重", "score", "index"]):
        return "多指标评价数据"
    if data_rows:
        return "表格型数据"
    return "题面数据/待补充附件"


def infer_output_type(problem_type: str, qtext: str) -> str:
    mapping = {
        "优化决策型": "最优方案、约束解释、敏感性分析",
        "数据评价型": "评价指标体系、权重、排序或等级",
        "预测回归型": "预测值、置信/预测区间、误差诊断",
        "机理仿真型": "仿真轨迹、参数解释、情景比较",
        "机器学习型": "分类/预测结果、性能指标、可解释性分析",
        "综合开放型": "综合模型链路、分问结论、最终建议",
    }
    return mapping.get(problem_type, "模型结果与结论解释")


def write_problem_statement(problem_text: str, problem_files: List[Dict[str, str]]) -> None:
    if problem_text.strip():
        content = "# 赛题原文（自动汇总）\n\n" + problem_text.strip() + "\n"
    else:
        content = "# 赛题原文（自动汇总）\n\n未读取到可解析题目文本。请把题目 txt/md/docx/pdf 放入 `00_problem/inbox/`。\n"
    OUT_PROBLEM.write_text(content, encoding="utf-8")


def write_attachments(problem_files: List[Dict[str, str]], data_rows: List[Dict[str, object]]) -> None:
    lines = ["# 附件总览（自动生成）", "", "## 题目文件", ""]
    if problem_files:
        lines.append("| 文件 | 类型 | 状态 | 抽取字符数 |")
        lines.append("|---|---:|---|---:|")
        for x in problem_files:
            lines.append(f"| {x['file']} | {x['ext']} | {x['status']} | {x['chars']} |")
    else:
        lines.append("未发现题目文件。")
    lines += ["", "## 数据文件", ""]
    if data_rows:
        lines.append("| 文件 | 类型 | 状态 | 规模/字段 |")
        lines.append("|---|---:|---|---|")
        for d in data_rows:
            scale = ""
            if "rows" in d:
                scale = f"{d.get('rows','')} 行 × {d.get('columns','')} 列"
            elif "sheets" in d:
                scale = "; ".join([f"{s.get('sheet')}:{s.get('rows')}x{s.get('columns')}" for s in d.get("sheets", [])])
            headers = d.get("headers") or d.get("keys") or []
            if headers:
                scale += "；字段：" + "、".join(map(str, list(headers)[:12]))
            lines.append(f"| {d.get('file','')} | {d.get('type','')} | {d.get('status','')} | {scale} |")
    else:
        lines.append("未发现数据文件。")
    OUT_ATTACH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_profile(questions: List[Tuple[str, str]], data_rows: List[Dict[str, object]]) -> List[Dict[str, str]]:
    hint = data_hints(data_rows)
    rows = []
    for q, text in questions:
        inf = score_rule(text, hint)
        rows.append({
            "question": q,
            "problem_type": inf["problem_type"],
            "model_family": inf["model_family"],
            "main_model": inf["main_model"],
            "baseline_model": inf["baseline_model"],
            "data_shape": infer_data_shape(data_rows, text),
            "output_type": infer_output_type(inf["problem_type"], text),
            "risk_focus": inf["risk_focus"],
            "confidence_0_1": inf["confidence_0_1"],
            "notes": inf["notes"],
        })
    with OUT_PROFILE.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["question", "problem_type", "model_family", "main_model", "baseline_model", "data_shape", "output_type", "risk_focus", "confidence_0_1", "notes"])
        w.writeheader()
        w.writerows(rows)
    return rows


def write_decomposition(questions: List[Tuple[str, str]], profile_rows: List[Dict[str, str]]) -> None:
    by_q = {r["question"]: r for r in profile_rows}
    lines = ["# 任务分解（投喂式启动自动生成）", "", "> 该文件是第一版自动判断结果，正式建模前必须人工复核。", ""]
    for q, text in questions:
        r = by_q.get(q, {})
        lines += [
            f"## {q}",
            "",
            f"- 原题片段：{text[:500]}",
            f"- 判断题型：{r.get('problem_type','')}",
            f"- 推荐模型族：{r.get('model_family','')}",
            f"- 候选主模型：{r.get('main_model','')}",
            f"- 基线模型：{r.get('baseline_model','')}",
            f"- 数据形态：{r.get('data_shape','')}",
            f"- 输出目标：{r.get('output_type','')}",
            f"- 主要风险：{r.get('risk_focus','')}",
            f"- 自动置信度：{r.get('confidence_0_1','')}",
            "",
            "### 下一步",
            "1. 核对题型和模型族是否符合赛题语义。",
            "2. 确认附件字段是否足以支撑该模型。",
            "3. 先建立基线模型，再引入复杂模型。",
            "4. 按 active_figure_plan.csv 先画结构图、关键结果图、检验图。",
            "",
        ]
    OUT_DECOMP.write_text("\n".join(lines), encoding="utf-8")


def write_data_dictionary(data_rows: List[Dict[str, object]]) -> None:
    fieldnames = ["file", "sheet", "field", "inferred_role", "sample_values", "notes"]
    rows = []
    for d in data_rows:
        if isinstance(d.get("sheets"), list):
            for s in d.get("sheets", []):
                for h in s.get("headers", [])[:80]:
                    rows.append({"file": d.get("file", ""), "sheet": s.get("sheet", ""), "field": h, "inferred_role": infer_field_role(str(h)), "sample_values": "", "notes": "自动字段识别，待人工补充含义和单位"})
        else:
            samples = d.get("samples", {}) if isinstance(d.get("samples"), dict) else {}
            headers = d.get("headers") or d.get("keys") or []
            for h in headers[:120]:
                rows.append({"file": d.get("file", ""), "sheet": "", "field": h, "inferred_role": infer_field_role(str(h)), "sample_values": "；".join(samples.get(h, [])[:3]) if isinstance(samples, dict) else "", "notes": "自动字段识别，待人工补充含义和单位"})
    with OUT_DATA_DICT.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)


def infer_field_role(field: str) -> str:
    f = field.lower()
    if any(k in f for k in ["id", "编号", "代码", "code"]):
        return "标识字段"
    if any(k in f for k in ["时间", "日期", "year", "month", "day", "time", "date"]):
        return "时间字段"
    if any(k in f for k in ["经度", "纬度", "lon", "lat", "lng", "坐标"]):
        return "空间字段"
    if any(k in f for k in ["类别", "类型", "class", "label", "标签"]):
        return "类别/标签字段"
    if any(k in f for k in ["评分", "得分", "score", "指标", "index"]):
        return "评价指标字段"
    if any(k in f for k in ["量", "率", "值", "price", "cost", "count", "number", "rate", "value"]):
        return "数值指标字段"
    return "待判断字段"


def write_reports(data_rows: List[Dict[str, object]], profile_rows: List[Dict[str, str]]) -> None:
    lines = ["# 数据质量初筛报告（自动生成）", ""]
    if not data_rows:
        lines.append("未发现可用数据文件。若赛题有附件，请放入 `03_data/raw/`。")
    else:
        for d in data_rows:
            lines.append(f"## {d.get('file','')}")
            lines.append(f"- 状态：{d.get('status','')}")
            if d.get("rows") != "":
                lines.append(f"- 规模：{d.get('rows','')} 行 × {d.get('columns','')} 列")
            if d.get("sheets"):
                lines.append("- 工作表：" + "；".join([f"{s.get('sheet')}({s.get('rows')}×{s.get('columns')})" for s in d.get("sheets", [])]))
            headers = d.get("headers") or d.get("keys") or []
            if headers:
                lines.append("- 字段预览：" + "、".join(map(str, headers[:20])))
            lines.append("")
    OUT_DATA_REPORT.write_text("\n".join(lines) + "\n", encoding="utf-8")

    low = [r for r in profile_rows if float(r.get("confidence_0_1", "0")) < 0.65]
    miss = ["# 缺失信息与人工复核清单", ""]
    if low:
        miss.append("## 低置信度题目")
        for r in low:
            miss.append(f"- {r['question']}：置信度 {r['confidence_0_1']}，建议人工复核 `{r['problem_type']} / {r['model_family']}`。")
        miss.append("")
    if not data_rows:
        miss.append("- 未发现数据附件，无法做字段级模型匹配。")
    miss.append("- 若题目是图片扫描版，脚本不会自动 OCR，请补充文字版题目。")
    miss.append("- 若 PDF 抽取文字乱码，请把题面复制为 txt/md 后重新运行。")
    OUT_MISSING.write_text("\n".join(miss) + "\n", encoding="utf-8")

    mr = ["# 模型路线初稿（自动生成）", ""]
    for r in profile_rows:
        mr += [
            f"## {r['question']}",
            f"- 题型：{r['problem_type']}",
            f"- 模型族：{r['model_family']}",
            f"- 主模型：{r['main_model']}",
            f"- 基线模型：{r['baseline_model']}",
            f"- 数据形态：{r['data_shape']}",
            f"- 风险控制：{r['risk_focus']}",
            "",
        ]
    OUT_MODEL_ROUTE.write_text("\n".join(mr), encoding="utf-8")


def run_router() -> bool:
    script = ROOT / "scripts" / "route_weight_config.py"
    if not script.exists():
        print("[WARN] route_weight_config.py not found; skip dynamic routing.")
        return False
    result = subprocess.run([sys.executable, str(script)], cwd=str(ROOT), text=True, capture_output=True)
    if result.stdout:
        print(result.stdout.strip())
    if result.stderr:
        print(result.stderr.strip())
    return result.returncode == 0


def write_summary(problem_files: List[Dict[str, str]], data_rows: List[Dict[str, object]], profile_rows: List[Dict[str, str]], routed: bool) -> None:
    counts = Counter(r["model_family"] for r in profile_rows)
    lines = [
        "# 投喂式启动摘要",
        "",
        f"- 题目文件数：{len(problem_files)}",
        f"- 数据文件数：{len(data_rows)}",
        f"- 自动识别问题数：{len(profile_rows)}",
        f"- 动态权重路由：{'已完成' if routed else '未完成'}",
        f"- 模型族分布：" + "；".join([f"{k}×{v}" for k, v in counts.items()]),
        "",
        "## 已生成文件",
        "- `00_problem/problem_statement.md`",
        "- `00_problem/attachments_overview.md`",
        "- `01_task_analysis/problem_model_profile.csv`",
        "- `01_task_analysis/task_decomposition.md`",
        "- `03_data/data_dictionary.csv`",
        "- `03_data/data_quality_report.md`",
        "- `05_model/model_route.md`",
        "- `09_paper/active_weight_config.csv`",
        "- `08_figures/active_figure_plan.csv`",
        "",
        "## 必须人工确认",
        "1. `problem_model_profile.csv` 的题型和模型族是否准确。",
        "2. `data_dictionary.csv` 的字段含义、单位和缺失值规则。",
        "3. `active_figure_plan.csv` 中正文主图是否与最终模型一致。",
    ]
    OUT_SUMMARY.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    problem_text, problem_files = collect_problem_files()
    data_rows = collect_data_files()
    questions = split_questions(problem_text)
    profile_rows = write_profile(questions, data_rows)
    write_problem_statement(problem_text, problem_files)
    write_attachments(problem_files, data_rows)
    write_decomposition(questions, profile_rows)
    write_data_dictionary(data_rows)
    write_reports(data_rows, profile_rows)
    routed = run_router()
    write_summary(problem_files, data_rows, profile_rows, routed)

    print("[OK] feed-start finished")
    print(f"[OK] problem files: {len(problem_files)}; data files: {len(data_rows)}; questions: {len(profile_rows)}")
    print("[NEXT] Review 01_task_analysis/auto_start_summary.md and problem_model_profile.csv")


if __name__ == "__main__":
    main()
    try:
        from workflow_utils import complete_stage
        complete_stage("intake", "已完成赛题与数据投喂解析。请完成数据文件和题目文本人工复核。")
    except Exception as exc:
        print(f"[WARN] workflow_state 未更新: {exc}")
